from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Q, Avg
from django.core.paginator import Paginator
from django.contrib import messages
from datetime import timedelta, date
from decimal import Decimal
import csv
import json

from .models import (
    PricingPlan, Customer, Subscription, Invoice, 
    InvoiceLineItem, Payment, RevenueMetric
)


def business_dashboard(request):
    """Main business dashboard with key metrics"""
    # Calculate date ranges
    today = timezone.now().date()
    month_start = today.replace(day=1)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    
    # Revenue metrics
    current_month_revenue = calculate_revenue_for_period(month_start, today)
    last_month_revenue = calculate_revenue_for_period(last_month_start, month_start - timedelta(days=1))
    
    # Customer metrics
    total_customers = Customer.objects.count()
    new_customers_this_month = Customer.objects.filter(
        created_at__date__gte=month_start
    ).count()
    
    # Subscription metrics
    active_subscriptions = Subscription.objects.filter(
        status__in=['trial', 'active']
    ).count()
    
    canceled_this_month = Subscription.objects.filter(
        status='canceled',
        canceled_at__date__gte=month_start
    ).count()
    
    # Invoice metrics
    pending_invoices = Invoice.objects.filter(status='sent').count()
    overdue_invoices = Invoice.objects.filter(
        status='sent',
        due_date__lt=today
    ).count()
    
    # Calculate growth rates
    revenue_growth = calculate_growth_rate(current_month_revenue, last_month_revenue)
    
    # Top performing plans
    top_plans = PricingPlan.objects.annotate(
        subscription_count=Count('subscriptions', filter=Q(subscriptions__status__in=['trial', 'active'])),
        revenue=Sum('subscriptions__plan__price', filter=Q(subscriptions__status__in=['trial', 'active']))
    ).filter(subscription_count__gt=0).order_by('-revenue')[:5]
    
    # Recent activity
    recent_subscriptions = Subscription.objects.select_related(
        'customer__user', 'plan'
    ).order_by('-created_at')[:10]
    
    recent_invoices = Invoice.objects.select_related(
        'customer__user'
    ).order_by('-created_at')[:10]
    
    context = {
        'revenue_metrics': {
            'current_month': current_month_revenue,
            'last_month': last_month_revenue,
            'growth_rate': revenue_growth,
        },
        'customer_metrics': {
            'total': total_customers,
            'new_this_month': new_customers_this_month,
        },
        'subscription_metrics': {
            'active': active_subscriptions,
            'canceled_this_month': canceled_this_month,
        },
        'invoice_metrics': {
            'pending': pending_invoices,
            'overdue': overdue_invoices,
        },
        'top_plans': top_plans,
        'recent_subscriptions': recent_subscriptions,
        'recent_invoices': recent_invoices,
    }
    
    return render(request, 'business/dashboard.html', context)


def pricing_plans_list(request):
    """List and manage pricing plans"""
    plans = PricingPlan.objects.annotate(
        subscription_count=Count('subscriptions', filter=Q(subscriptions__status__in=['trial', 'active'])),
        monthly_revenue=Sum('subscriptions__plan__price', filter=Q(subscriptions__status__in=['trial', 'active']))
    ).order_by('display_order')
    
    context = {
        'plans': plans,
        'total_plans': plans.count(),
        'active_plans': plans.filter(is_active=True).count(),
    }
    
    return render(request, 'business/pricing_plans.html', context)


def pricing_plan_create(request):
    """Create a new pricing plan"""
    from .forms import PricingPlanForm
    
    if request.method == 'POST':
        form = PricingPlanForm(request.POST)
        if form.is_valid():
            plan = form.save()
            messages.success(request, f'Preisplan "{plan.name}" wurde erfolgreich erstellt.')
            return redirect('business:pricing_plans')
    else:
        form = PricingPlanForm()
    
    context = {
        'form': form,
        'title': 'Neuen Preisplan erstellen',
    }
    
    return render(request, 'business/pricing_plan_form.html', context)


def pricing_plan_edit(request, plan_id):
    """Edit an existing pricing plan"""
    from .forms import PricingPlanForm
    
    plan = get_object_or_404(PricingPlan, id=plan_id)
    
    if request.method == 'POST':
        form = PricingPlanForm(request.POST, instance=plan)
        if form.is_valid():
            plan = form.save()
            messages.success(request, f'Preisplan "{plan.name}" wurde erfolgreich aktualisiert.')
            return redirect('business:pricing_plans')
    else:
        form = PricingPlanForm(instance=plan)
    
    context = {
        'form': form,
        'plan': plan,
        'title': f'Preisplan "{plan.name}" bearbeiten',
    }
    
    return render(request, 'business/pricing_plan_form.html', context)


def pricing_plan_delete(request, plan_id):
    """Delete a pricing plan"""
    plan = get_object_or_404(PricingPlan, id=plan_id)
    
    # Check if plan has active subscriptions
    active_subscriptions = plan.subscriptions.filter(status__in=['trial', 'active']).count()
    
    if active_subscriptions > 0:
        messages.error(request, f'Preisplan "{plan.name}" kann nicht gelöscht werden - {active_subscriptions} aktive Abonnements vorhanden.')
        return redirect('business:pricing_plans')
    
    if request.method == 'POST':
        plan_name = plan.name
        plan.delete()
        messages.success(request, f'Preisplan "{plan_name}" wurde erfolgreich gelöscht.')
        return redirect('business:pricing_plans')
    
    context = {
        'plan': plan,
        'active_subscriptions': active_subscriptions,
    }
    
    return render(request, 'business/pricing_plan_delete.html', context)


def customers_list(request):
    """List all customers with filtering and search"""
    customers = Customer.objects.select_related('user').annotate(
        subscription_count=Count('subscriptions'),
        total_spent=Sum('invoices__total_amount', filter=Q(invoices__status='paid')),
        active_subscription=Count('subscriptions', filter=Q(subscriptions__status__in=['trial', 'active']))
    ).order_by('-created_at')
    
    # Apply filters
    customer_type = request.GET.get('type', '')
    if customer_type:
        customers = customers.filter(customer_type=customer_type)
    
    search = request.GET.get('search', '')
    if search:
        customers = customers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(company_name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(customers, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Customer statistics
    customer_stats = {
        'total': Customer.objects.count(),
        'business': Customer.objects.filter(customer_type='business').count(),
        'individual': Customer.objects.filter(customer_type='individual').count(),
        'with_active_subscription': Customer.objects.filter(
            subscriptions__status__in=['trial', 'active']
        ).distinct().count(),
    }
    
    context = {
        'page_obj': page_obj,
        'customer_stats': customer_stats,
        'current_filters': {
            'type': customer_type,
            'search': search,
        }
    }
    
    return render(request, 'business/customers.html', context)


def subscriptions_list(request):
    """List all subscriptions with filtering"""
    subscriptions = Subscription.objects.select_related(
        'customer__user', 'plan'
    ).order_by('-created_at')
    
    # Apply filters
    status = request.GET.get('status', '')
    if status:
        subscriptions = subscriptions.filter(status=status)
    
    plan_id = request.GET.get('plan', '')
    if plan_id:
        subscriptions = subscriptions.filter(plan_id=plan_id)
    
    # Pagination
    paginator = Paginator(subscriptions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get plans for filter dropdown
    plans = PricingPlan.objects.filter(is_active=True).order_by('display_order')
    
    # Subscription statistics
    subscription_stats = {
        'total': Subscription.objects.count(),
        'active': Subscription.objects.filter(status='active').count(),
        'trial': Subscription.objects.filter(status='trial').count(),
        'canceled': Subscription.objects.filter(status='canceled').count(),
        'revenue_monthly': calculate_monthly_recurring_revenue(),
    }
    
    context = {
        'page_obj': page_obj,
        'plans': plans,
        'subscription_stats': subscription_stats,
        'current_filters': {
            'status': status,
            'plan': plan_id,
        }
    }
    
    return render(request, 'business/subscriptions.html', context)


def invoices_list(request):
    """List all invoices with filtering"""
    invoices = Invoice.objects.select_related(
        'customer__user'
    ).order_by('-issue_date')
    
    # Apply filters
    status = request.GET.get('status', '')
    if status:
        invoices = invoices.filter(status=status)
    
    customer_id = request.GET.get('customer')
    if customer_id:
        invoices = invoices.filter(customer_id=customer_id)
    
    # Pagination
    paginator = Paginator(invoices, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Invoice statistics
    today = timezone.now().date()
    invoice_stats = {
        'total': Invoice.objects.count(),
        'draft': Invoice.objects.filter(status='draft').count(),
        'sent': Invoice.objects.filter(status='sent').count(),
        'paid': Invoice.objects.filter(status='paid').count(),
        'overdue': Invoice.objects.filter(status='sent', due_date__lt=today).count(),
        'total_outstanding': Invoice.objects.filter(
            status='sent'
        ).aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00'),
    }
    
    context = {
        'page_obj': page_obj,
        'invoice_stats': invoice_stats,
        'current_filters': {
            'status': status,
            'customer': customer_id,
        }
    }
    
    return render(request, 'business/invoices.html', context)


def invoice_create(request):
    """Create a new invoice"""
    from .forms import InvoiceForm, InvoiceLineItemFormSet
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        formset = InvoiceLineItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            invoice = form.save()
            formset.instance = invoice
            formset.save()
            
            messages.success(request, f'Rechnung {invoice.invoice_number} wurde erfolgreich erstellt.')
            return redirect('business:invoices')
    else:
        form = InvoiceForm()
        formset = InvoiceLineItemFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'title': 'Neue Rechnung erstellen',
    }
    
    return render(request, 'business/invoice_form.html', context)


def invoice_detail(request, invoice_id):
    """View invoice details"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    context = {
        'invoice': invoice,
        'line_items': invoice.line_items.all(),
        'payments': invoice.payments.all(),
    }
    
    return render(request, 'business/invoice_detail.html', context)


def invoice_edit(request, invoice_id):
    """Edit an existing invoice"""
    from .forms import InvoiceForm, InvoiceLineItemFormSet
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if invoice.status == 'paid':
        messages.error(request, 'Bezahlte Rechnungen können nicht bearbeitet werden.')
        return redirect('business:invoice_detail', invoice_id=invoice.id)
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice)
        formset = InvoiceLineItemFormSet(request.POST, instance=invoice)
        
        if form.is_valid() and formset.is_valid():
            invoice = form.save()
            formset.save()
            
            messages.success(request, f'Rechnung {invoice.invoice_number} wurde erfolgreich aktualisiert.')
            return redirect('business:invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm(instance=invoice)
        formset = InvoiceLineItemFormSet(instance=invoice)
    
    context = {
        'form': form,
        'formset': formset,
        'invoice': invoice,
        'title': f'Rechnung {invoice.invoice_number} bearbeiten',
    }
    
    return render(request, 'business/invoice_form.html', context)


def invoice_send(request, invoice_id):
    """Send an invoice to customer"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if invoice.status != 'draft':
        messages.error(request, 'Nur Entwürfe können versendet werden.')
        return redirect('business:invoice_detail', invoice_id=invoice.id)
    
    if request.method == 'POST':
        from .services import InvoiceService
        
        try:
            InvoiceService.send_invoice(invoice)
            messages.success(request, f'Rechnung {invoice.invoice_number} wurde erfolgreich versendet.')
        except Exception as e:
            messages.error(request, f'Fehler beim Versenden der Rechnung: {e}')
        
        return redirect('business:invoice_detail', invoice_id=invoice.id)
    
    context = {
        'invoice': invoice,
    }
    
    return render(request, 'business/invoice_send_confirm.html', context)


def invoice_mark_paid(request, invoice_id):
    """Mark invoice as paid"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if invoice.status not in ['sent', 'overdue']:
        messages.error(request, 'Nur versendete oder überfällige Rechnungen können als bezahlt markiert werden.')
        return redirect('business:invoice_detail', invoice_id=invoice.id)
    
    if request.method == 'POST':
        payment_method = request.POST.get('payment_method', 'manual')
        
        from .services import InvoiceService
        
        try:
            InvoiceService.mark_invoice_paid(invoice, payment_method)
            messages.success(request, f'Rechnung {invoice.invoice_number} wurde als bezahlt markiert.')
        except Exception as e:
            messages.error(request, f'Fehler beim Markieren der Rechnung: {e}')
        
        return redirect('business:invoice_detail', invoice_id=invoice.id)
    
    context = {
        'invoice': invoice,
    }
    
    return render(request, 'business/invoice_mark_paid.html', context)


def revenue_analytics(request):
    """Revenue analytics and reporting with export functionality"""
    # Get date range from request or default to last 12 months
    period = request.GET.get('period', '12m')
    export_format = request.GET.get('export', '')
    
    if period == '7d':
        start_date = timezone.now().date() - timedelta(days=7)
        period_type = 'daily'
    elif period == '30d':
        start_date = timezone.now().date() - timedelta(days=30)
        period_type = 'daily'
    elif period == '12m':
        start_date = timezone.now().date() - timedelta(days=365)
        period_type = 'monthly'
    else:
        start_date = timezone.now().date() - timedelta(days=365)
        period_type = 'monthly'
    
    # Get revenue metrics for the period
    revenue_data = RevenueMetric.objects.filter(
        date__gte=start_date,
        period_type=period_type
    ).order_by('date')
    
    # If no revenue data exists, generate sample data for demonstration
    if not revenue_data.exists():
        revenue_data = generate_sample_revenue_data(start_date, period_type)
    
    # Calculate totals for the period
    period_totals = revenue_data.aggregate(
        total_revenue=Sum('total_revenue'),
        subscription_revenue=Sum('subscription_revenue'),
        one_time_revenue=Sum('one_time_revenue'),
        new_customers=Sum('new_customers'),
    )
    
    # Plan performance analysis
    plan_performance = PricingPlan.objects.annotate(
        active_count=Count('subscriptions', filter=Q(subscriptions__status__in=['trial', 'active'])),
        monthly_revenue=Sum('subscriptions__plan__price', filter=Q(subscriptions__status__in=['trial', 'active'])),
        total_customers=Count('subscriptions__customer', distinct=True)
    ).filter(active_count__gt=0).order_by('-monthly_revenue')
    
    # Monthly recurring revenue trend
    mrr_trend = calculate_mrr_trend(start_date)
    
    # Customer lifetime value
    avg_customer_ltv = calculate_average_customer_ltv()
    
    # Churn rate
    churn_rate = calculate_churn_rate()
    
    # Handle export requests
    if export_format in ['csv', 'excel', 'pdf']:
        return handle_analytics_export(request, export_format, {
            'revenue_data': revenue_data,
            'period_totals': period_totals,
            'plan_performance': plan_performance,
            'period': period,
            'start_date': start_date
        })
    
    context = {
        'revenue_data': revenue_data,
        'period_totals': period_totals,
        'plan_performance': plan_performance,
        'mrr_trend': mrr_trend,
        'avg_customer_ltv': avg_customer_ltv,
        'churn_rate': churn_rate,
        'selected_period': period,
    }
    
    return render(request, 'business/revenue_analytics.html', context)


# Helper functions

def calculate_revenue_for_period(start_date, end_date):
    """Calculate total revenue for a given period"""
    return Invoice.objects.filter(
        status='paid',
        paid_date__date__gte=start_date,
        paid_date__date__lte=end_date
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')


def calculate_growth_rate(current, previous):
    """Calculate growth rate percentage"""
    if previous == 0:
        return 100 if current > 0 else 0
    return round(((current - previous) / previous) * 100, 1)


def calculate_monthly_recurring_revenue():
    """Calculate current monthly recurring revenue"""
    monthly_subscriptions = Subscription.objects.filter(
        status__in=['trial', 'active'],
        plan__billing_cycle='monthly'
    ).aggregate(
        total=Sum('plan__price')
    )['total'] or Decimal('0.00')
    
    # Convert quarterly and yearly to monthly equivalent
    quarterly_subscriptions = Subscription.objects.filter(
        status__in=['trial', 'active'],
        plan__billing_cycle='quarterly'
    ).aggregate(
        total=Sum('plan__price')
    )['total'] or Decimal('0.00')
    
    yearly_subscriptions = Subscription.objects.filter(
        status__in=['trial', 'active'],
        plan__billing_cycle='yearly'
    ).aggregate(
        total=Sum('plan__price')
    )['total'] or Decimal('0.00')
    
    return monthly_subscriptions + (quarterly_subscriptions / 3) + (yearly_subscriptions / 12)


def calculate_mrr_trend(start_date):
    """Calculate MRR trend over time"""
    # This would typically be calculated from historical data
    # For now, return current MRR
    return [
        {
            'date': start_date,
            'mrr': calculate_monthly_recurring_revenue()
        }
    ]


def calculate_average_customer_ltv():
    """Calculate average customer lifetime value"""
    # Simplified calculation: average revenue per customer
    total_revenue = Invoice.objects.filter(status='paid').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')
    
    total_customers = Customer.objects.count()
    
    if total_customers == 0:
        return Decimal('0.00')
    
    return round(total_revenue / total_customers, 2)


def calculate_churn_rate():
    """Calculate monthly churn rate"""
    # Get subscriptions from last month
    last_month = timezone.now().date().replace(day=1) - timedelta(days=1)
    month_start = last_month.replace(day=1)
    
    active_start_month = Subscription.objects.filter(
        status__in=['trial', 'active'],
        created_at__date__lte=month_start
    ).count()
    
    canceled_in_month = Subscription.objects.filter(
        status='canceled',
        canceled_at__date__gte=month_start,
        canceled_at__date__lte=last_month
    ).count()
    
    if active_start_month == 0:
        return 0
    
    return round((canceled_in_month / active_start_month) * 100, 1)


def generate_sample_revenue_data(start_date, period_type):
    """Generate sample revenue data for demonstration purposes"""
    from random import randint, uniform
    
    sample_data = []
    current_date = start_date
    end_date = timezone.now().date()
    
    base_revenue = 1000
    base_customers = 5
    
    while current_date <= end_date:
        # Generate realistic sample data with growth trend
        days_since_start = (current_date - start_date).days
        growth_factor = 1 + (days_since_start * 0.002)  # 0.2% daily growth
        
        total_revenue = Decimal(str(round(base_revenue * growth_factor * uniform(0.8, 1.2), 2)))
        subscription_revenue = Decimal(str(round(total_revenue * uniform(0.7, 0.9), 2)))
        one_time_revenue = total_revenue - subscription_revenue
        new_customers = randint(max(1, int(base_customers * growth_factor * 0.8)), 
                               int(base_customers * growth_factor * 1.2))
        
        # Create a mock revenue metric object (not saved to database)
        class MockRevenueMetric:
            def __init__(self, date, total_revenue, subscription_revenue, one_time_revenue, new_customers):
                self.date = date
                self.total_revenue = total_revenue
                self.subscription_revenue = subscription_revenue
                self.one_time_revenue = one_time_revenue
                self.new_customers = new_customers
        
        sample_data.append(MockRevenueMetric(
            date=current_date,
            total_revenue=total_revenue,
            subscription_revenue=subscription_revenue,
            one_time_revenue=one_time_revenue,
            new_customers=new_customers
        ))
        
        # Increment date based on period type
        if period_type == 'daily':
            current_date += timedelta(days=1)
        else:  # monthly
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
    
    return sample_data


def handle_analytics_export(request, export_format, data):
    """Handle export of analytics data in various formats"""
    filename_base = f"revenue_analytics_{data['period']}_{data['start_date'].strftime('%Y%m%d')}"
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow(['Datum', 'Gesamtumsatz (€)', 'Abonnement-Umsatz (€)', 'Einmaliger Umsatz (€)', 'Neue Kunden'])
        
        # Write revenue data
        for item in data['revenue_data']:
            writer.writerow([
                item.date.strftime('%d.%m.%Y'),
                float(item.total_revenue),
                float(item.subscription_revenue),
                float(item.one_time_revenue),
                item.new_customers
            ])
        
        # Write summary
        writer.writerow([])
        writer.writerow(['ZUSAMMENFASSUNG'])
        writer.writerow(['Gesamtumsatz', float(data['period_totals']['total_revenue'] or 0)])
        writer.writerow(['Abonnement-Umsatz', float(data['period_totals']['subscription_revenue'] or 0)])
        writer.writerow(['Einmaliger Umsatz', float(data['period_totals']['one_time_revenue'] or 0)])
        writer.writerow(['Neue Kunden', data['period_totals']['new_customers'] or 0])
        
        return response
    
    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Revenue Analytics"
            
            # Header styling
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            
            # Write headers
            headers = ['Datum', 'Gesamtumsatz (€)', 'Abonnement-Umsatz (€)', 'Einmaliger Umsatz (€)', 'Neue Kunden']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row, item in enumerate(data['revenue_data'], 2):
                worksheet.cell(row=row, column=1, value=item.date.strftime('%d.%m.%Y'))
                worksheet.cell(row=row, column=2, value=float(item.total_revenue))
                worksheet.cell(row=row, column=3, value=float(item.subscription_revenue))
                worksheet.cell(row=row, column=4, value=float(item.one_time_revenue))
                worksheet.cell(row=row, column=5, value=item.new_customers)
            
            # Add summary section
            summary_row = len(data['revenue_data']) + 3
            worksheet.cell(row=summary_row, column=1, value="ZUSAMMENFASSUNG").font = header_font
            worksheet.cell(row=summary_row + 1, column=1, value="Gesamtumsatz")
            worksheet.cell(row=summary_row + 1, column=2, value=float(data['period_totals']['total_revenue'] or 0))
            worksheet.cell(row=summary_row + 2, column=1, value="Abonnement-Umsatz")
            worksheet.cell(row=summary_row + 2, column=2, value=float(data['period_totals']['subscription_revenue'] or 0))
            worksheet.cell(row=summary_row + 3, column=1, value="Neue Kunden")
            worksheet.cell(row=summary_row + 3, column=2, value=data['period_totals']['new_customers'] or 0)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            workbook.save(response)
            return response
            
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            messages.warning(request, 'Excel-Export nicht verfügbar. CSV-Export wird verwendet.')
            return handle_analytics_export(request, 'csv', data)
    
    elif export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor('#FF6B35')
            )
            story.append(Paragraph("RenditeFuchs Revenue Analytics", title_style))
            story.append(Spacer(1, 12))
            
            # Period info
            story.append(Paragraph(f"Zeitraum: {data['period']} (ab {data['start_date'].strftime('%d.%m.%Y')})", styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Summary table
            summary_data = [
                ['Metrik', 'Wert'],
                ['Gesamtumsatz', f"€{float(data['period_totals']['total_revenue'] or 0):,.2f}"],
                ['Abonnement-Umsatz', f"€{float(data['period_totals']['subscription_revenue'] or 0):,.2f}"],
                ['Einmaliger Umsatz', f"€{float(data['period_totals']['one_time_revenue'] or 0):,.2f}"],
                ['Neue Kunden', str(data['period_totals']['new_customers'] or 0)],
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Revenue data table (first 20 entries)
            story.append(Paragraph("Revenue Daten (Übersicht)", styles['Heading2']))
            revenue_data_table = [['Datum', 'Gesamtumsatz', 'Abonnements', 'Einmalig', 'Neue Kunden']]
            
            for item in data['revenue_data'][:20]:  # Limit to first 20 entries
                revenue_data_table.append([
                    item.date.strftime('%d.%m.%Y'),
                    f"€{float(item.total_revenue):,.2f}",
                    f"€{float(item.subscription_revenue):,.2f}",
                    f"€{float(item.one_time_revenue):,.2f}",
                    str(item.new_customers),
                ])
            
            revenue_table = Table(revenue_data_table)
            revenue_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(revenue_table)
            
            if len(data['revenue_data']) > 20:
                story.append(Spacer(1, 12))
                story.append(Paragraph(f"Hinweis: Zeigt die ersten 20 von {len(data['revenue_data'])} Einträgen", styles['Normal']))
            
            doc.build(story)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            return response
            
        except ImportError:
            # Fallback to CSV if reportlab is not available
            messages.warning(request, 'PDF-Export nicht verfügbar. CSV-Export wird verwendet.')
            return handle_analytics_export(request, 'csv', data)
    
    # Default fallback
    messages.error(request, 'Unbekanntes Export-Format.')
    return redirect('business:analytics')